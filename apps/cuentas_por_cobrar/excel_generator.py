"""
Generador de Excel del estado de cuenta de un cliente (CxC).
apps/cuentas_por_cobrar/excel_generator.py

Hojas: Resumen, Cuentas, Cuotas, Abonos (openpyxl).
"""
from io import BytesIO

from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

AZUL = '1E40AF'
MONEDA = '#,##0.00'


def _encabezar(ws, titulos):
    ws.append(titulos)
    for celda in ws[1]:
        celda.font = Font(bold=True, color='FFFFFF')
        celda.fill = PatternFill('solid', fgColor=AZUL)
        celda.alignment = Alignment(horizontal='center')


def _autoancho(ws):
    for columna in ws.columns:
        ancho = max((len(str(c.value)) for c in columna if c.value is not None), default=10)
        ws.column_dimensions[columna[0].column_letter].width = min(ancho + 3, 40)


def generar_estado_cuenta_xlsx(cliente, cuentas, resumen) -> BytesIO:
    cuentas = list(cuentas)
    wb = Workbook()

    ws = wb.active
    ws.title = 'Resumen'
    ws.append(['Estado de cuenta', cliente.nombre])
    ws.append(['Cedula/RNC', cliente.cedula_rnc or '-'])
    ws.append(['Generado', timezone.localtime().strftime('%d/%m/%Y %H:%M')])
    ws.append([])
    ws.append(['Limite de credito', float(resumen['limite_credito'])])
    ws.append(['Saldo pendiente', float(resumen['saldo_pendiente'])])
    ws.append(['Credito disponible', float(resumen['credito_disponible'])])
    ws.append(['Monto vencido', float(resumen['monto_vencido'])])
    ws.append([
        'Proximo vencimiento',
        resumen['proximo_vencimiento'].strftime('%d/%m/%Y') if resumen['proximo_vencimiento'] else '-',
    ])
    ws['A1'].font = Font(bold=True, size=14, color=AZUL)
    for fila in range(5, 9):
        ws.cell(row=fila, column=2).number_format = MONEDA
    _autoancho(ws)

    ws = wb.create_sheet('Cuentas')
    _encabezar(ws, ['Venta', 'Emision', 'Vence', 'Capital', 'Interes %', 'Interes', 'Total financiado', 'Saldo', 'Estado'])
    for cuenta in cuentas:
        ws.append([
            cuenta.venta.numero_venta,
            cuenta.fecha_emision.strftime('%d/%m/%Y'),
            cuenta.fecha_limite.strftime('%d/%m/%Y'),
            float(cuenta.saldo_original),
            float(cuenta.interes_porcentaje),
            float(cuenta.monto_interes),
            float(cuenta.monto_financiado),
            float(cuenta.saldo),
            cuenta.estado,
        ])
    for fila in ws.iter_rows(min_row=2):
        for celda in (fila[3], fila[5], fila[6], fila[7]):
            celda.number_format = MONEDA
    _autoancho(ws)

    ws = wb.create_sheet('Cuotas')
    _encabezar(ws, ['Venta', 'Cuota', 'Vence', 'Monto', 'Saldo', 'Estado'])
    for cuenta in cuentas:
        for cuota in cuenta.cuotas.all():
            ws.append([
                cuenta.venta.numero_venta,
                cuota.numero,
                cuota.fecha_vencimiento.strftime('%d/%m/%Y'),
                float(cuota.monto),
                float(cuota.saldo),
                cuota.estado,
            ])
    for fila in ws.iter_rows(min_row=2):
        for celda in (fila[3], fila[4]):
            celda.number_format = MONEDA
    _autoancho(ws)

    ws = wb.create_sheet('Abonos')
    _encabezar(ws, ['Fecha', 'Venta', 'Metodo', 'Referencia', 'Monto', 'Estado', 'Motivo anulacion'])
    abonos = [
        (cuenta, pago)
        for cuenta in cuentas
        for pago in cuenta.pagos_cxc.all()
    ]
    abonos.sort(key=lambda par: par[1].fecha_pago, reverse=True)
    for cuenta, pago in abonos:
        ws.append([
            timezone.localtime(pago.fecha_pago).strftime('%d/%m/%Y %H:%M'),
            cuenta.venta.numero_venta,
            pago.metodo,
            pago.referencia or '-',
            float(pago.monto),
            pago.estado,
            pago.motivo_anulacion or '',
        ])
    for fila in ws.iter_rows(min_row=2):
        fila[4].number_format = MONEDA
    _autoancho(ws)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
